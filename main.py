import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
import pandas as pd
import time
import openpyxl

options = uc.ChromeOptions()

# Remove the line below since we don't need to detach the browser window
# options.add_experimental_option("detach", True)

# Load Credentials
excel_file_path = "all.xlsx"
df = pd.read_excel(excel_file_path)

scraped_data = []

# Create excel
output_file_path = "file.xlsx"
wb = openpyxl.Workbook()
sheet = wb.active

# Write header for the columns
sheet.cell(row=1, column=1, value="Username")
sheet.cell(row=1, column=2, value="Scraped Data")

# Iterate through the rows of the DataFrame and input credentials
for row_number, row in df.iterrows():
    username = row["Username"]
    password = row["Password"]

    browser = uc.Chrome()

    browser.get("URL")
    browser.maximize_window()

    time.sleep(7)

    # Locate the username and password fields and input the data
    username_field = browser.find_element(By.ID, "username-textbox")
    password_field = browser.find_element(By.ID, "password-textbox")

    username_field.send_keys(username)
    time.sleep(7)
    password_field.send_keys(password)
    time.sleep(7)

    # Submit the form (you may need to adjust the submit method based on your specific webpage)
    browser.find_element(By.ID, "login-button").click()

    time.sleep(60)
    # Data scraping.
    try:
        plan_names = browser.find_element(By.XPATH, "/html/body/div/main/div[2]/div/div[2]/section/div[2]/div[1]/div/div[1]/div[2]").text
        scraped_data.append(plan_names)
    except Exception as e:
        print(f"Error occurred while scraping data for user '{username}': {e}")
        scraped_data.append("Failed")

    # Write the data after scraping all users
    for idx, (username, data) in enumerate(zip(df["Username"], scraped_data)):
        sheet.cell(row=idx+2, column=1, value=username)
        sheet.cell(row=idx+2, column=2, value=data)

    wb.save("plan_names.xlsx")


    # Close the current browser instance
    browser.quit()

    # Add some delay to avoid overloading the server (optional)
    time.sleep(30)


# Save excel

print("Data scraped and saved to excel")
